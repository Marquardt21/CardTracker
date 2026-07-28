"""
Set import service.

Two import paths behind one URL field:
  * an Upper Deck checklist page (HTML table) — scraped with BeautifulSoup
  * a Beckett-style .xlsx/.xls checklist file — parsed with openpyxl

After any import, run reconciliation against unmatched cards in the collection.
"""
import logging
import re
from datetime import datetime
from io import BytesIO
from urllib.parse import unquote, urlparse

import httpx
from bs4 import BeautifulSoup
from sqlalchemy.orm import Session

from backend.config import DEFAULT_SPORT, SCRAPER_HEADERS, SPORTS
from backend.models import Card, SetChecklist, SetChecklistCard
from backend.schemas import ReconciliationResult

logger = logging.getLogger(__name__)


async def fetch_and_parse_url(url: str) -> dict | None:
    """Parse a checklist URL with whichever reader matches it.

    Returns the same dict shape from both paths:
    {"set_name", "brand", "year", "sport", "cards": [...]}, or None on failure.
    """
    if _is_workbook_url(url):
        return await fetch_xlsx_url(url)
    return await scrape_upper_deck_url(url)


def _is_workbook_url(url: str) -> bool:
    return urlparse(url).path.lower().endswith((".xlsx", ".xls"))


# ---------------------------------------------------------------------------
# Upper Deck URL scraper
# ---------------------------------------------------------------------------

async def scrape_upper_deck_url(url: str) -> dict | None:
    """
    Fetch an Upper Deck checklist page and parse it into a structured dict.
    Returns None on any failure so the caller can show a friendly error.
    """
    try:
        async with httpx.AsyncClient(timeout=15.0, follow_redirects=True, headers=SCRAPER_HEADERS) as client:
            resp = await client.get(url)
            if resp.status_code != 200:
                logger.warning("Upper Deck URL returned %s: %s", resp.status_code, url)
                return None
            return _parse_upper_deck_html(resp.text, url)
    except httpx.TimeoutException:
        logger.warning("Timeout fetching Upper Deck URL: %s", url)
        return None
    except Exception as exc:
        logger.warning("Error fetching Upper Deck URL %s: %s", url, exc)
        return None


def _parse_upper_deck_html(html: str, url: str) -> dict | None:
    """
    Parse the Upper Deck checklist HTML.
    Columns: Set Name, Card #, Description, Team City, Team Name,
             Rookie, Auto, Tech, #'d, SPs, Stated Odds, Point
    """
    soup = BeautifulSoup(html, "html.parser")

    # Derive set name and year from page title or URL
    set_name, brand, year = _extract_set_metadata(soup, url)

    # Find the checklist table — UD uses a table with these header columns
    table = _find_checklist_table(soup)
    if table is None:
        logger.warning("No checklist table found at %s", url)
        return None

    headers, data_start = _extract_table_headers(table)
    col = _column_index(headers)

    cards = []
    for row in table.find_all("tr")[data_start:]:
        cells = row.find_all("td")
        if not cells:
            continue
        card = _parse_row(cells, col, set_name)
        if card:
            cards.append(card)

    if not cards:
        logger.warning("Table found but no cards parsed at %s", url)
        return None

    # Upper Deck's checklists are hockey; the preview screen lets the user change it.
    return {"set_name": set_name, "brand": brand, "year": year, "sport": "Hockey", "cards": cards}


def _extract_set_metadata(soup: BeautifulSoup, url: str) -> tuple[str, str, int]:
    brand = "Upper Deck"
    year = datetime.utcnow().year

    title = soup.find("h1") or soup.find("title")
    title_text = title.get_text(strip=True) if title else ""

    # Pull year from title or URL
    year_match = re.search(r"(20\d{2})", title_text + url)
    if year_match:
        year = int(year_match.group(1))

    set_name = title_text or _set_name_from_url(url)
    # Clean common suffixes
    set_name = re.sub(r"\s*(checklist|hockey cards?|cards?)\s*$", "", set_name, flags=re.I).strip()
    if not set_name:
        set_name = _set_name_from_url(url)

    return set_name, brand, year


def _set_name_from_url(url: str) -> str:
    slug = url.rstrip("/").split("/")[-1]
    return slug.replace("-", " ").replace("_", " ").title()


def _extract_table_headers(table) -> tuple[list[str], int]:
    """
    Return (header_list, first_data_row_index).

    Some UD pages (e.g. O-Pee-Chee Platinum) put an empty <th> row first and
    place the real column names in the first <td> row.  Fall back to that row
    when all <th> text is empty.
    """
    th_texts = [th.get_text(strip=True).lower() for th in table.find_all("th")]
    if any(th_texts):
        return th_texts, 1

    # Fallback: find the first <td> row whose text matches known column names
    known = {"card", "card #", "description", "player", "set name"}
    for i, row in enumerate(table.find_all("tr")):
        tds = row.find_all("td")
        texts = [td.get_text(strip=True).lower() for td in tds]
        if known & set(texts):          # at least one known header present
            return texts, i + 1         # data starts on the next row
    return [], 1


def _find_checklist_table(soup: BeautifulSoup):
    # Upper Deck wraps the checklist in a table; look for one with "Card #" header
    for table in soup.find_all("table"):
        text = table.get_text(separator=" ", strip=True).lower()
        if "card #" in text or "card#" in text or "description" in text:
            return table
    # Fallback: largest table on the page
    tables = soup.find_all("table")
    return max(tables, key=lambda t: len(t.find_all("tr"))) if tables else None


def _column_index(headers: list[str]) -> dict[str, int]:
    """Map expected column names to their indices (flexible matching)."""
    mapping = {}
    for i, h in enumerate(headers):
        if "set name" in h or "set" == h:
            mapping.setdefault("set_variant", i)
        elif ("card" in h and "#" in h) or h == "card":
            mapping["card_number"] = i
        elif "team city" in h or "city" in h:
            mapping.setdefault("team_city", i)
        elif "team name" in h or "team" == h:
            mapping.setdefault("team_name", i)
        elif "description" in h or "player" in h or h == "name":
            mapping["player_name"] = i
        elif "rookie" in h:
            mapping["rookie"] = i
        elif "auto" in h:
            mapping["auto"] = i
        elif "tech" in h or "relic" in h or "patch" in h:
            mapping["relic"] = i
        elif "#'d" in h or "numbered" in h or "print" in h:
            mapping["print_run"] = i
    return mapping


def _card_type_from_flags(is_auto: bool, is_relic: bool, parallel_color: str | None, is_rookie: bool) -> str:
    """Card type priority, shared by both import paths: auto/relic take priority,
    then parallel (any parallel_color), then rookie, then base."""
    if is_auto and is_relic:
        return "patch_relic"
    if is_auto:
        return "autograph"
    if parallel_color:
        return "parallel"
    if is_rookie:
        return "rookie"
    return "base"


def _parse_row(cells: list, col: dict, set_name: str) -> dict | None:
    def cell(key, default=""):
        idx = col.get(key)
        if idx is None or idx >= len(cells):
            return default
        return cells[idx].get_text(strip=True)

    card_number = cell("card_number").lstrip("#")
    player_name = cell("player_name")

    if not card_number or not player_name or player_name.lower() in ("description", "player"):
        return None

    # Team
    team_city = cell("team_city")
    team_name = cell("team_name")
    team = f"{team_city} {team_name}".strip() if team_city or team_name else None

    # Compute parallel_color first — any non-base variant name is a parallel
    variant = cell("set_variant")
    base_labels = {set_name.lower(), "base", "base set", "base set - rookies", ""}
    parallel_color = variant if (variant and variant.lower() not in base_labels) else None

    is_rookie = cell("rookie") not in ("", "0", "No", "N")
    is_auto = cell("auto") not in ("", "0", "No", "N")
    is_relic = cell("relic") not in ("", "0", "No", "N")

    card_type = _card_type_from_flags(is_auto, is_relic, parallel_color, is_rookie)

    # Print run
    print_run_raw = cell("print_run")
    print_run = None
    if print_run_raw:
        m = re.search(r"(\d+)", print_run_raw)
        if m:
            print_run = int(m.group(1))

    return {
        "card_number": card_number,
        "player_name": player_name,
        "is_rookie": is_rookie,
        "has_auto": is_auto,
        "is_serialed": print_run is not None,
        "team": team,
        "card_type": card_type,
        "parallel_color": parallel_color,
        "print_run": print_run,
    }


# ---------------------------------------------------------------------------
# .xlsx checklist reader (Beckett-hosted workbooks)
# ---------------------------------------------------------------------------

_TEAM_SETS_SHEET = "team sets"
_XLSX_HEADER_LABELS = {"card #", "card#", "card number", "card no", "#"}
_BRAND_KEYWORDS = [
    "Upper Deck", "Topps", "Bowman", "Panini", "Donruss", "Leaf", "Fleer", "Score",
]


async def fetch_xlsx_url(url: str) -> dict | None:
    """Download an .xlsx/.xls checklist and parse its "Team Sets" sheet."""
    try:
        async with httpx.AsyncClient(timeout=30.0, follow_redirects=True, headers=SCRAPER_HEADERS) as client:
            resp = await client.get(url)
            if resp.status_code != 200:
                logger.warning("Checklist file returned %s: %s", resp.status_code, url)
                return None
            return parse_xlsx(resp.content, url)
    except httpx.TimeoutException:
        logger.warning("Timeout fetching checklist file: %s", url)
        return None
    except Exception as exc:
        logger.warning("Error fetching checklist file %s: %s", url, exc)
        return None


def parse_xlsx(data: bytes, url: str) -> dict | None:
    """Parse a checklist workbook.

    The "Team Sets" sheet holds one row per card:
    [Subset Name] | [Card #] | [Player] | [Team] | [RC flag]
    """
    from openpyxl import load_workbook

    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        logger.warning("Could not open workbook %s: %s", url, exc)
        return None

    sheet = next((wb[n] for n in wb.sheetnames if n.strip().lower() == _TEAM_SETS_SHEET), None)
    if sheet is None:
        logger.warning("No 'Team Sets' sheet in %s (sheets: %s)", url, wb.sheetnames)
        return None

    set_name, brand, year, sport = _metadata_from_filename(url)

    cards = []
    for row in sheet.iter_rows(values_only=True):
        card = _parse_xlsx_row(row, set_name)
        if card:
            cards.append(card)

    if not cards:
        logger.warning("'Team Sets' sheet found but no cards parsed at %s", url)
        return None

    return {"set_name": set_name, "brand": brand, "year": year, "sport": sport, "cards": cards}


def _parse_xlsx_row(row: tuple, set_name: str) -> dict | None:
    def cell(i: int) -> str:
        return str(row[i]).strip() if i < len(row) and row[i] is not None else ""

    subset, card_number, player_name = cell(0), cell(1).lstrip("#"), cell(2)
    if not card_number or not player_name:
        return None
    if card_number.lower() in _XLSX_HEADER_LABELS or player_name.lower() in ("player", "player name", "description"):
        return None

    team = cell(3) or None
    rc_flag = cell(4)

    card_type, parallel_color, is_rookie, is_auto = _classify_subset(subset, set_name)
    is_rookie = is_rookie or rc_flag.lower() in ("rc", "yes", "y", "true", "1", "rookie")
    if card_type == "base" and is_rookie:
        card_type = "rookie"

    # Print run isn't carried on this sheet — left null (nullable field).
    return {
        "card_number": card_number,
        "player_name": player_name,
        "is_rookie": is_rookie,
        "has_auto": is_auto,
        "is_serialed": False,
        "team": team,
        "card_type": card_type,
        "parallel_color": parallel_color,
        "print_run": None,
    }


def _classify_subset(subset: str, set_name: str) -> tuple[str, str | None, bool, bool]:
    """Derive (card_type, parallel_color, is_rookie, has_auto) from a subset name.

    Same keyword priority as the Upper Deck parser, but every signal comes from
    the one subset-name column instead of dedicated flag columns.
    """
    subset = (subset or "").strip()
    low = subset.lower()

    is_auto   = "auto" in low or "signature" in low
    is_relic  = any(k in low for k in ("relic", "patch", "jersey", "memorabilia"))
    is_rookie = bool(re.search(r"\b(rc|rookie|rookies)\b", low))

    base_labels = {set_name.lower(), "base", "base set", "base set - rookies", ""}
    is_base = low in base_labels
    # A relic/auto subset names the hit, not a colour parallel.
    parallel_color = None if (is_base or is_auto or is_relic) else (subset or None)

    # patch_relic covers both auto+relic and relic-only subsets; everything else
    # follows the shared priority chain.
    card_type = "patch_relic" if is_relic else _card_type_from_flags(is_auto, False, parallel_color, is_rookie)
    return card_type, parallel_color, is_rookie, is_auto


def _metadata_from_filename(url: str) -> tuple[str, str, int, str]:
    """Best-effort set name / brand / year / sport from the file name.

    Guesses are shown editable on the import preview screen, so being wrong here
    is a correction the user makes rather than a failed import.
    """
    name = unquote(urlparse(url).path.rstrip("/").split("/")[-1])
    name = re.sub(r"\.xlsx?$", "", name, flags=re.I)
    words = " ".join(re.sub(r"[-_]+", " ", name).split())
    low = words.lower()

    year_match = re.search(r"(19|20)\d{2}", words)
    year = int(year_match.group(0)) if year_match else datetime.utcnow().year
    sport = next((s for s in SPORTS if s.lower() in low), DEFAULT_SPORT)
    brand = next((b for b in _BRAND_KEYWORDS if b.lower() in low), "")

    set_name = re.sub(r"\b(19|20)\d{2}(-\d{2})?\b", " ", words)
    set_name = re.sub(
        r"\b(checklist|cards?|" + "|".join(SPORTS) + r")\b", " ", set_name, flags=re.I
    )
    set_name = " ".join(set_name.split()) or words

    return set_name, brand or set_name.split(" ")[0], year, sport


# ---------------------------------------------------------------------------
# Save to DB + reconciliation
# ---------------------------------------------------------------------------

def save_set_and_reconcile(
    db: Session, parsed: dict, source_url: str
) -> tuple[SetChecklist, ReconciliationResult]:
    checklist = SetChecklist(
        set_name=parsed["set_name"],
        brand=parsed["brand"],
        year=parsed["year"],
        sport=parsed.get("sport") or DEFAULT_SPORT,
        total_cards=len(parsed["cards"]),
        source_url=source_url,
        imported_at=datetime.utcnow(),
    )
    db.add(checklist)
    db.flush()

    for c in parsed["cards"]:
        db.add(SetChecklistCard(
            set_id=checklist.id,
            card_number=c["card_number"],
            player_name=c["player_name"],
            team=c.get("team"),
            card_type=c.get("card_type", "base"),
            parallel_color=c.get("parallel_color"),
            print_run=c.get("print_run"),
            is_serialed=c.get("is_serialed", False),
            has_auto=c.get("has_auto", False),
            is_rookie=c.get("is_rookie", False),
        ))

    db.commit()
    db.refresh(checklist)
    reconciliation = reconcile_unmatched(db)
    return checklist, reconciliation


def reconcile_unmatched(db: Session) -> ReconciliationResult:
    """Match unmatched collection cards against all imported checklists."""
    unmatched = db.query(Card).filter(Card.checklist_matched == False).all()  # noqa: E712
    newly_matched = 0

    for card in unmatched:
        if match_card_to_checklists(db, card):
            newly_matched += 1

    still_unmatched_cards = db.query(Card).filter(Card.checklist_matched == False).all()  # noqa: E712
    return ReconciliationResult(
        newly_matched=newly_matched,
        still_unmatched=len(still_unmatched_cards),
        unmatched_cards=still_unmatched_cards,
    )


def match_card_to_checklists(db: Session, card: Card) -> bool:
    """
    Try to find a matching set_checklist_card for a collection card.
    Matches on card_number + fuzzy player name within sets that match brand/year.
    Returns True if a match was found and saved.
    """
    candidates = (
        db.query(SetChecklistCard)
        .join(SetChecklist)
        .filter(
            SetChecklistCard.card_number == card.card_number,
            SetChecklist.brand.ilike(f"%{card.brand}%"),
        )
        .all()
    )

    for candidate in candidates:
        if _names_match(card.player_name, candidate.player_name):
            candidate.owned = True
            candidate.collection_card_id = card.id
            card.checklist_matched = True
            db.commit()
            return True

    return False


def _names_match(a: str, b: str) -> bool:
    """Fuzzy name match: exact after lowercasing, or last name match."""
    a, b = a.lower().strip(), b.lower().strip()
    if a == b:
        return True
    # Last name match
    return a.split()[-1] == b.split()[-1]
